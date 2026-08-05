#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitHub Stars Exporter v2.3
导出 GitHub Stars 为 JSON / CSV / XLSX / HTML
"""

import requests
import json
import csv
import base64
import os
import sys
import time
import argparse
import html as html_module
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

API_BASE = "https://api.github.com"


def get_headers(token):
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    return headers


def fetch_all_starred(username, token):
    repos = []
    page = 1
    per_page = 100
    headers = get_headers(token)
    print(f"[1/4] 正在获取用户 {username} 的星标仓库...")
    while True:
        url = f"{API_BASE}/users/{username}/starred?per_page={per_page}&page={page}"
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 404:
            print(f"  错误：用户 '{username}' 不存在")
            return []
        if resp.status_code == 403:
            print(f"  错误：API 速率限制。建议添加 GITHUB_TOKEN")
            return []
        resp.raise_for_status()
        data = resp.json()
        if not data:
            break
        repos.extend(data)
        print(f"  第 {page} 页: {len(data)} 个，累计 {len(repos)} 个")
        if len(data) < per_page:
            break
        page += 1
        time.sleep(0.2)
    print(f"  共获取 {len(repos)} 个星标仓库\n")
    return repos


def fetch_readme(owner, repo_name, token):
    url = f"{API_BASE}/repos/{owner}/{repo_name}/readme"
    try:
        resp = requests.get(url, headers=get_headers(token), timeout=15)
        if resp.status_code == 404:
            return ""
        resp.raise_for_status()
        data = resp.json()
        return base64.b64decode(data.get("content", "")).decode("utf-8", errors="ignore")
    except Exception:
        return ""


def build_repo_data(repos, token, fetch_readme_flag=True):
    results = []
    total = len(repos)
    print(f"[2/4] 正在处理 {total} 个仓库...")
    for idx, repo in enumerate(repos, 1):
        owner = repo["owner"]["login"]
        name = repo["name"]
        full_name = repo["full_name"]
        readme = fetch_readme(owner, name, token) if fetch_readme_flag else ""
        if fetch_readme_flag:
            time.sleep(0.08)
        item = {
            "序号": idx, "项目名称": name, "项目全名": full_name,
            "项目链接": repo["html_url"], "仓库描述": repo.get("description") or "",
            "主页网址": repo.get("homepage") or "", "README": readme,
            "项目语言": repo.get("language") or "", "星标数": repo.get("stargazers_count", 0),
            "Fork数": repo.get("forks_count", 0), "Watch数": repo.get("watchers_count", 0),
            "Open_Issues": repo.get("open_issues_count", 0), "默认分支": repo.get("default_branch", ""),
            "创建时间": repo.get("created_at", ""), "更新时间": repo.get("updated_at", ""),
            "推送时间": repo.get("pushed_at", ""), "仓库大小_KB": repo.get("size", 0),
            "是否为Fork": repo.get("fork", False), "Topics": ", ".join(repo.get("topics", [])),
            "License": (repo.get("license") or {}).get("name", ""), "Owner": owner,
            "Owner类型": repo["owner"].get("type", ""), "Owner头像": repo["owner"].get("avatar_url", ""),
        }
        results.append(item)
        if idx % 200 == 0 or idx == total:
            print(f"  [{idx}/{total}] 已处理")
    print(f"  全部处理完成\n")
    return results


def export_json(data, filepath):
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON: {filepath}")


def export_csv(data, filepath):
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    fieldnames = list(data[0].keys())
    csv_fieldnames = [f for f in fieldnames if f not in ["README", "Owner头像"]]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow({k: v for k, v in row.items() if k in csv_fieldnames})
    print(f"  CSV: {filepath}")


def export_xlsx(data, filepath):
    if not HAS_OPENPYXL:
        print("  未安装 openpyxl，跳过 XLSX")
        return
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GitHub Stars"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="24292F", end_color="24292F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )
    fieldnames = list(data[0].keys())
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, row in enumerate(data, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = cell_align
            cell.border = thin_border
    col_widths = {
        "序号": 6, "项目名称": 25, "项目全名": 35, "项目链接": 45,
        "仓库描述": 50, "主页网址": 40, "README": 60, "项目语言": 12,
        "星标数": 10, "Fork数": 10, "Watch数": 10, "Open_Issues": 12,
        "默认分支": 12, "创建时间": 20, "更新时间": 20, "推送时间": 20,
        "仓库大小_KB": 14, "是否为Fork": 12, "Topics": 40, "License": 20,
        "Owner": 18, "Owner类型": 12, "Owner头像": 45
    }
    for col_idx, field in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"  XLSX: {filepath}")


def export_html(data, filepath):
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)

    rows_html = ""
    headers = list(data[0].keys())
    display_headers = [h for h in headers if h not in ["README", "Owner头像"]]

    for row in data:
        cells = ""
        for h in display_headers:
            val = row.get(h, "")
            if h == "项目链接" and val:
                cells += f''<td><a href="{html_module.escape(str(val))}" target="_blank">{html_module.escape(str(val))}</a></td>''
            elif h == "主页网址" and val:
                cells += f''<td><a href="{html_module.escape(str(val))}" target="_blank">{html_module.escape(str(val))}</a></td>''
            else:
                cells += f"<td>{html_module.escape(str(val))}</td>"
        rows_html += f"<tr>{cells}</tr>\n"

    header_cells = "".join(f"<th>{html_module.escape(h)}</th>" for h in display_headers)

    # 语言筛选：项目语言在第7列（索引6）
    html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GitHub Stars</title>
<style>
  :root {{ --bg: #f6f8fa; --fg: #24292f; --border: #d0d7de; --header-bg: #24292f; --header-fg: #fff; --link: #0969da; }}
  @media (prefers-color-scheme: dark) {{ :root {{ --bg: #0d1117; --fg: #c9d1d9; --border: #30363d; --header-bg: #161b22; --header-fg: #c9d1d9; --link: #58a6ff; }} }}
  body {{ font-family: -apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif; background: var(--bg); color: var(--fg); margin: 0; padding: 20px; }}
  h1 {{ margin-bottom: 10px; }}
  .info {{ color: #57606a; margin-bottom: 20px; font-size: 14px; }}
  .controls {{ margin-bottom: 15px; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  input, select {{ padding: 6px 10px; border: 1px solid var(--border); border-radius: 6px; background: var(--bg); color: var(--fg); font-size: 14px; }}
  .download-links {{ margin-bottom: 15px; }}
  .download-links a {{ display: inline-block; margin-right: 12px; color: var(--link); text-decoration: none; font-size: 14px; }}
  .download-links a:hover {{ text-decoration: underline; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th, td {{ border: 1px solid var(--border); padding: 8px; text-align: left; vertical-align: top; }}
  th {{ background: var(--header-bg); color: var(--header-fg); position: sticky; top: 0; }}
  tr:hover {{ background: rgba(128,128,128,0.05); }}
  a {{ color: var(--link); text-decoration: none; }}
</style>
</head>
<body>
<h1>⭐ GitHub Stars</h1>
<div class="info">共 {len(data)} 个仓库，导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
<div class="download-links">
  <a href="stars.json" download>📥 下载 JSON</a>
  <a href="stars.csv" download>📥 下载 CSV</a>
  <a href="stars.xlsx" download>📥 下载 XLSX</a>
</div>
<div class="controls">
  <input type="text" id="search" placeholder="搜索项目名/描述/语言..." oninput="filterTable()">
  <select id="langFilter" onchange="filterTable()"><option value="">全部语言</option></select>
</div>
<table id="starTable">
<thead><tr>{header_cells}</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<script>
const rows = Array.from(document.querySelectorAll('#starTable tbody tr'));
const langs = [...new Set(rows.map(r => r.cells[6]?.textContent).filter(Boolean))].sort();
const sel = document.getElementById('langFilter');
langs.forEach(l => {{ const o=document.createElement('option'); o.value=l; o.textContent=l; sel.appendChild(o); }});
function filterTable() {{
  const q = document.getElementById('search').value.toLowerCase();
  const lang = document.getElementById('langFilter').value;
  rows.forEach(r => {{
    const text = r.textContent.toLowerCase();
    const rowLang = r.cells[6]?.textContent || '';
    r.style.display = (text.includes(q) && (!lang || rowLang === lang)) ? '' : 'none';
  }});
}}
</script>
</body>
</html>'''

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"  HTML: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="导出 GitHub Stars")
    parser.add_argument("--output", "-o", default="./dist", help="输出目录 (默认: ./dist)")
    parser.add_argument("--username", "-u", default=os.environ.get("GITHUB_USERNAME"), help="GitHub 用户名")
    parser.add_argument("--token", "-t", default=os.environ.get("GITHUB_TOKEN"), help="GitHub Token")
    parser.add_argument("--formats", "-f", default="json,csv,xlsx,html", help="导出格式，逗号分隔")
    parser.add_argument("--no-readme", action="store_true", help="不获取 README 内容")
    args = parser.parse_args()

    username = args.username
    token = args.token
    output_dir = args.output
    formats = [f.strip().lower() for f in args.formats.split(",")]
    fetch_readme_flag = not args.no_readme

    if not username:
        print("错误：未指定 GitHub 用户名。请通过 --username 参数或 GITHUB_USERNAME 环境变量设置。")
        sys.exit(1)

    print(f"输出目录: {output_dir}")
    print(f"导出格式: {formats}")
    print(f"获取 README: {fetch_readme_flag}\n")

    repos = fetch_all_starred(username, token)
    if not repos:
        print("未获取到任何星标仓库，退出。")
        sys.exit(1)

    data = build_repo_data(repos, token, fetch_readme_flag)

    print("[3/4] 正在导出文件...")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"github_stars_{username}_{timestamp}"

    for fmt in formats:
        if fmt == "json":
            export_json(data, os.path.join(output_dir, f"{base_name}.json"))
            export_json(data, os.path.join(output_dir, "stars.json"))
        elif fmt == "csv":
            export_csv(data, os.path.join(output_dir, f"{base_name}.csv"))
            export_csv(data, os.path.join(output_dir, "stars.csv"))
        elif fmt == "xlsx":
            export_xlsx(data, os.path.join(output_dir, f"{base_name}.xlsx"))
            export_xlsx(data, os.path.join(output_dir, "stars.xlsx"))
        elif fmt == "html":
            export_html(data, os.path.join(output_dir, "index.html"))
        else:
            print(f"  未知格式: {fmt}")

    print("\n[4/4] 全部完成！")
    print(f"  输出目录: {output_dir}")
    for f in sorted(os.listdir(output_dir)):
        print(f"    - {f}")


if __name__ == "__main__":
    main()
